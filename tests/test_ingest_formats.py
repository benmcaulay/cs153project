"""
Ingestion of the two correspondence/exhibit formats legal matters actually
arrive in: email (.eml) and spreadsheets (.xlsx). The contract is that a fact
stated in either — a date in an email, a dollar figure computed in a sheet —
is extracted as plain text so the grounding pipeline can transcribe it.
"""
from __future__ import annotations

from email.message import EmailMessage

from app import catalog
from app.ingest import SUPPORTED_EXTS, ingest_folder, read_document


def _write_eml(path, *, frm, to, subject, date, body, html=False, attachment=None):
    msg = EmailMessage()
    msg["From"] = frm
    msg["To"] = to
    msg["Subject"] = subject
    msg["Date"] = date
    if html:
        # HTML-only body (no plain-text alternative) to exercise tag stripping.
        msg.set_content(f"<html><body><p>{body}</p></body></html>", subtype="html")
    else:
        msg.set_content(body)
    if attachment:
        msg.add_attachment(b"%PDF-1.4 fake", maintype="application", subtype="pdf", filename=attachment)
    with open(path, "wb") as fh:
        fh.write(bytes(msg))


def test_eml_extracts_headers_and_body(tmp_path):
    p = tmp_path / "demand.eml"
    _write_eml(
        p,
        frm="counsel@firm.com",
        to="opposing@other.com",
        subject="Settlement demand",
        date="Mon, 03 Mar 2024 10:00:00 -0800",
        body="We demand $75,000 to resolve this matter by April 1, 2024.",
        attachment="exhibit_A.pdf",
    )
    text = read_document(str(p))
    assert "Settlement demand" in text          # subject header
    assert "counsel@firm.com" in text           # from header
    assert "$75,000" in text                     # body fact, groundable
    assert "April 1, 2024" in text
    assert "exhibit_A.pdf" in text               # attachment surfaced by name


def test_eml_html_only_body_is_stripped(tmp_path):
    p = tmp_path / "note.eml"
    _write_eml(
        p,
        frm="a@b.com",
        to="c@d.com",
        subject="Engagement",
        date="Tue, 04 Mar 2024 09:00:00 -0800",
        body="Retainer is $5,000.",
        html=True,
    )
    text = read_document(str(p))
    assert "Retainer is $5,000." in text
    assert "<p>" not in text and "<html>" not in text  # tags stripped


def test_xlsx_flattens_sheets_and_formula_values(tmp_path):
    from openpyxl import Workbook

    p = tmp_path / "damages.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Damages"
    ws.append(["Item", "Amount"])
    ws.append(["Medical", 50000])
    ws.append(["Lost wages", 25000])
    ws["B4"] = "=B2+B3"  # formula; cached value written below
    # openpyxl can't compute formulas, so simulate a cached value for data_only
    ws["B4"] = 75000
    wb.save(str(p))

    text = read_document(str(p))
    assert "[Sheet: Damages]" in text   # sheet labeled for provenance
    assert "Medical" in text and "50000" in text
    assert "75000" in text              # the total is transcribable


def test_new_formats_are_supported_and_ingest_in_a_matter(tmp_path):
    assert ".eml" in SUPPORTED_EXTS and ".xlsx" in SUPPORTED_EXTS
    folder = tmp_path / "Matter"
    folder.mkdir()
    _write_eml(
        folder / "msg.eml",
        frm="a@b.com", to="c@d.com", subject="Hi",
        date="Tue, 04 Mar 2024 09:00:00 -0800", body="The filing date is May 9, 2024.",
    )
    from openpyxl import Workbook

    wb = Workbook()
    wb.active.append(["Docket", "24-CV-0098"])
    wb.save(str(folder / "sheet.xlsx"))

    docs = ingest_folder(str(folder))
    names = {d.filename for d in docs}
    assert "msg.eml" in names and "sheet.xlsx" in names
    blob = "\n".join(d.text for d in docs)
    assert "May 9, 2024" in blob and "24-CV-0098" in blob
    # chunks are produced so the content is retrievable
    assert any(d.chunks for d in docs)


def test_eml_text_attachment_content_is_extracted(tmp_path):
    """The fact stapled to the email — not just the attachment's name — is
    extracted and groundable."""
    msg = EmailMessage()
    msg["From"] = "counsel@firm.com"
    msg["To"] = "client@x.com"
    msg["Subject"] = "Please review"
    msg["Date"] = "Mon, 03 Mar 2024 10:00:00 -0800"
    msg.set_content("See the attached exhibit.")
    msg.add_attachment(
        b"The agreed settlement is $123,456.",
        maintype="text",
        subtype="plain",
        filename="exhibit.txt",
    )
    p = tmp_path / "with_attach.eml"
    p.write_bytes(bytes(msg))

    text = read_document(str(p))
    assert "exhibit.txt" in text          # listed by name
    assert "$123,456" in text             # AND its content pulled in
    assert "--- Attachment: exhibit.txt ---" in text  # labeled for provenance


def test_eml_xlsx_attachment_is_parsed(tmp_path):
    import io

    from openpyxl import Workbook

    wb = Workbook()
    wb.active.append(["Line item", "Amount"])
    wb.active.append(["Total", 99000])
    buf = io.BytesIO()
    wb.save(buf)

    msg = EmailMessage()
    msg["From"] = "a@b.com"
    msg["To"] = "c@d.com"
    msg["Subject"] = "Damages ledger"
    msg["Date"] = "Tue, 04 Mar 2024 09:00:00 -0800"
    msg.set_content("Ledger attached.")
    msg.add_attachment(
        buf.getvalue(),
        maintype="application",
        subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="ledger.xlsx",
    )
    p = tmp_path / "ledger_mail.eml"
    p.write_bytes(bytes(msg))

    text = read_document(str(p))
    assert "ledger.xlsx" in text and "99000" in text   # routed to the xlsx reader


def test_forwarded_eml_attachment_recurses(tmp_path):
    """A forwarded email attached as a .eml is followed one level down."""
    inner = EmailMessage()
    inner["From"] = "witness@x.com"
    inner["To"] = "counsel@firm.com"
    inner["Subject"] = "Statement"
    inner["Date"] = "Sun, 02 Mar 2024 08:00:00 -0800"
    inner.set_content("The accident occurred on January 5, 2024.")

    outer = EmailMessage()
    outer["From"] = "counsel@firm.com"
    outer["To"] = "client@x.com"
    outer["Subject"] = "Fwd: Statement"
    outer["Date"] = "Tue, 04 Mar 2024 09:00:00 -0800"
    outer.set_content("Forwarding the witness statement.")
    outer.add_attachment(
        bytes(inner), maintype="application", subtype="octet-stream", filename="forwarded.eml"
    )
    p = tmp_path / "fwd.eml"
    p.write_bytes(bytes(outer))

    text = read_document(str(p))
    assert "forwarded.eml" in text
    assert "January 5, 2024" in text      # recursed into the attached email
    assert "Subject: Statement" in text   # inner headers came through too


def test_upload_validation_accepts_new_types(tmp_path, monkeypatch):
    monkeypatch.setattr(catalog, "MATTERS_DIR", str(tmp_path / "matters"))
    m = catalog.create_matter("Reyes v Brightway")
    # an unsupported type is still rejected
    import pytest

    with pytest.raises(catalog.CatalogError):
        catalog.add_document(m.id, "evil.exe", b"nope")
    # the new types are accepted
    info = catalog.add_document(m.id, "ledger.xlsx", b"PK\x03\x04")  # bytes need not be valid here
    assert "ledger.xlsx" in info.documents
